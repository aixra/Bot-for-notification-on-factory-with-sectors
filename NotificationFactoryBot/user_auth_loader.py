from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

import config


def load_allowed_logins() -> set[str]:
    path = Path(config.USERS_EXCEL_FILE)
    if not path.exists():
        return set()

    workbook = load_workbook(path, data_only=True)
    sheet_name = config.USERS_SHEET_NAME
    sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active

    allowed: set[str] = set()
    for row in sheet.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        raw_login = str(row[0]).strip()
        if not raw_login or raw_login.lower() == "login":
            continue
        normalized = raw_login.lstrip("@").lower()
        if normalized:
            allowed.add(normalized)

    return allowed


def is_login_allowed(login: str) -> bool:
    normalized = login.strip().lstrip("@").lower()
    if not normalized:
        return False
    return normalized in load_allowed_logins()
