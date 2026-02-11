from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

import config


def load_catalog() -> dict[str, list[str]]:
    path = Path(config.EQUIPMENT_EXCEL_FILE)
    if not path.exists():
        return {k: list(v) for k, v in config.DEFAULT_DEVICES.items()}

    workbook = load_workbook(path, data_only=True)
    sheet_name = config.EQUIPMENT_SHEET_NAME
    sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active

    catalog: dict[str, list[str]] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        # Первый столбец — сектор
        sector = str(row[0]).strip() if row[0] is not None else ""
        if not sector:
            continue

        # Создаём пустой список для сектора, если его ещё нет
        catalog.setdefault(sector, [])

        # Пробегаем все остальные колонки после первой
        for device in row[1:]:
            if device is None:
                continue
            device_name = str(device).strip()
            if device_name and device_name not in catalog[sector]:
                catalog[sector].append(device_name)

    return catalog if catalog else {k: list(v) for k, v in config.DEFAULT_DEVICES.items()}
